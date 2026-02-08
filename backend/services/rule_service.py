"""
Rule Service - Business Logic Layer
====================================
규칙 파일 업로드, 다운로드, 조회 등의 비즈니스 로직 처리
"""

import io
import json
from typing import List, Dict, Any, Optional
from uuid import UUID, uuid4
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database.rule_repository import RuleRepository
from utils.excel_parser import parse_rules_from_excel
from models import RuleFileUpload, RuleFileResponse, RuleCreate


class RuleService:
    """
    Service layer for rule management
    """

    def __init__(self, repository=None, ai_cache_service=None):
        """Initialize service with repository and cache service"""
        self.repository = repository or RuleRepository()
        self.ai_cache_service = ai_cache_service

    async def upload_rule_file(
        self,
        excel_content: bytes,
        metadata: RuleFileUpload
    ) -> RuleFileResponse:
        """
        규칙 파일을 업로드하고 해석함 (Smart Interpret 및 엑셀 내 기해석 정보 활용)
        """
        print(f"[RuleService] Starting upload for file: {metadata.file_name}")

        try:
            # Step 1: Parse Excel file
            natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row = parse_rules_from_excel(excel_content)

            # Step 2: Semantic Deduplication using centralized AICacheService
            print("[RuleService] Preparing rules for batch insert (Deduplication)...")
            rules_to_insert = []
            seen_rules = set()
            duplicate_count = 0
            
            for rule in natural_language_rules:
                field_name = rule["field"].strip() if rule["field"] else ""
                rule_text = rule["rule_text"].strip() if rule["rule_text"] else ""
                condition = rule["condition"].strip() if rule["condition"] else ""
                prefilled = rule.get("prefilled_ai", {})
                
                # Default AI interpretation data
                ai_rule_id = prefilled.get("ai_rule_id")
                ai_rule_type = prefilled.get("ai_rule_type")
                ai_parameters = prefilled.get("ai_parameters")
                ai_summary = prefilled.get("ai_interpretation_summary")
                ai_error = prefilled.get("ai_error_message")
                ai_confidence = 1.0 if ai_rule_type else None
                ai_model = "excel-import" if ai_rule_type else None
                is_common = prefilled.get("is_common", False)
                
                # 만약 엑셀에 AI 해석 정보가 없고 스마트 해석 엔진이 있다면 실행
                if not ai_rule_type and self.ai_cache_service:
                    interpreted, source = await self.ai_cache_service.smart_interpret_single(rule_text, field_name)
                    ai_rule_id = interpreted.get('rule_id', f"RULE_{uuid4().hex[:8]}")
                    ai_rule_type = interpreted.get('rule_type')
                    ai_parameters = interpreted.get('parameters')
                    ai_summary = interpreted.get('interpretation_summary')
                    ai_error = interpreted.get('error_message')
                    ai_confidence = float(interpreted.get('confidence_score', 0.8))
                    ai_model = f"smart-{source}"
                
                rule_params_key = json.dumps(ai_parameters, sort_keys=True) if ai_parameters else "{}"
                rule_key = (field_name, ai_rule_type, rule_params_key, condition)
                
                if rule_key in seen_rules:
                    duplicate_count += 1
                    continue
                
                seen_rules.add(rule_key)
                
                # DB 레코드 생성
                rule_record = {
                    "sheet_name": "Common",
                    "row_number": rule["row"],
                    "column_letter": str(rule["column_letter"])[:10] if rule["column_letter"] else "",
                    "field_name": rule["field"],
                    "rule_text": rule["rule_text"],
                    "condition": rule["condition"],
                    "note": rule["note"],
                    "is_active": True,
                    "is_common": is_common,
                    "ai_rule_id": ai_rule_id,
                    "ai_rule_type": ai_rule_type,
                    "ai_parameters": ai_parameters,
                    "ai_error_message": ai_error,
                    "ai_interpretation_summary": ai_summary,
                    "ai_confidence_score": ai_confidence,
                    "ai_model_version": ai_model
                }
                rules_to_insert.append(rule_record)

            # Step 3: Check for duplicate file and auto-increment version
            final_version = metadata.file_version or "1.0"
            existing_files = await self.repository.list_rule_files(status='active', limit=100)
            for existing in existing_files:
                if existing['file_name'] == metadata.file_name and existing['file_version'] == final_version:
                    try:
                        final_version = f"{float(final_version) + 0.1:.1f}"
                    except: final_version += "_new"

            # Step 4: Create rule_file record
            file_data = {
                "file_name": metadata.file_name,
                "file_version": final_version,
                "uploaded_by": metadata.uploaded_by or "system",
                "total_rules_count": len(natural_language_rules),
                "sheet_count": len(sheet_row_counts),
                "notes": metadata.notes,
                "status": "active"
            }
            file_id = await self.repository.create_rule_file(file_data)

            # Step 5: Batch insert rules
            for r in rules_to_insert: r["rule_file_id"] = file_id
            await self.repository.create_rules_batch(rules_to_insert)

            # Step 6: Save original file for future re-interpretation
            await self.repository.save_original_file(UUID(file_id), excel_content)

            # Step 7: Automatic interpretation for those NOT pre-filled
            if self.ai_cache_service:
                await self.ai_cache_service.interpret_and_cache_rules(file_id)

            file_record = await self.repository.get_rule_file(UUID(file_id))
            return RuleFileResponse(
                id=file_record['id'],
                file_name=file_record['file_name'],
                file_version=file_record.get('file_version'),
                uploaded_by=file_record.get('uploaded_by'),
                uploaded_at=file_record['uploaded_at'],
                sheet_count=file_record['sheet_count'],
                total_rules_count=file_record['total_rules_count'],
                status=file_record['status']
            )

        except Exception as e:
            print(f"[RuleService] Upload failed: {e}")
            raise Exception(f"Failed to upload rule file: {e}")

    async def export_rules_to_excel(self, file_id: str) -> bytes:
        """규칙을 Excel로 내보내기 (Parser와 호환되는 헤더 사용)"""
        all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "규칙 목록"
        
        # Parser의 markers와 정확히 일치하는 헤더
        headers = [
            "번호", "컬럼", "필드명", "규칙 내용", "조건", "비고", "공통 여부",
            "AI 해석 여부", "AI 규칙 ID", "AI 규칙 유형", "AI 파라미터(JSON)", 
            "AI 신뢰도", "AI 에러 메시지", "AI 해석 요약"
        ]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for i, r in enumerate(all_rules, 2): # Data starts from row 2 (Sub-header logic skipped for simplicity here)
            ws.cell(row=i, column=1, value=i-1)
            ws.cell(row=i, column=2, value=r.get('column_letter'))
            ws.cell(row=i, column=3, value=r.get('field_name'))
            ws.cell(row=i, column=4, value=r.get('rule_text'))
            ws.cell(row=i, column=5, value=r.get('condition'))
            ws.cell(row=i, column=6, value=r.get('note'))
            ws.cell(row=i, column=7, value="예" if r.get('is_common') else "아니오")
            ws.cell(row=i, column=8, value="예" if r.get('ai_rule_id') else "아니오")
            ws.cell(row=i, column=9, value=r.get('ai_rule_id'))
            ws.cell(row=i, column=10, value=r.get('ai_rule_type'))
            ws.cell(row=i, column=11, value=json.dumps(r.get('ai_parameters'), ensure_ascii=False) if r.get('ai_parameters') else "")
            ws.cell(row=i, column=12, value=r.get('ai_confidence_score'))
            ws.cell(row=i, column=13, value=r.get('ai_error_message'))
            ws.cell(row=i, column=14, value=r.get('ai_interpretation_summary'))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # --- 기타 메서드 (기존 유지) ---
    async def get_rule_file_details(self, file_id: str):
        """
        규칙 파일의 상세 정보 및 포함된 모든 규칙 조회
        """
        file_record = await self.repository.get_rule_file(UUID(file_id))
        if not file_record:
            return None
            
        all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)
        
        # 시트별로 규칙 그룹화
        from collections import defaultdict
        sheet_groups = defaultdict(list)
        for r in all_rules:
            sheet_name = r.get('sheet_name') or 'Common'
            sheet_groups[sheet_name].append({
                "id": str(r['id']),
                "column_name": r.get('field_name'),
                "rule_text": r.get('rule_text'),
                "ai_rule_type": r.get('ai_rule_type'),
                "ai_confidence_score": r.get('ai_confidence_score', 0) or 0
            })
            
        sheets_list = []
        for name, rules in sheet_groups.items():
            sheets_list.append({
                "sheet_name": name,
                "rule_count": len(rules),
                "sample_rules": rules
            })
            
        return {
            "id": str(file_record['id']),
            "file_name": file_record['file_name'],
            "file_version": file_record['file_version'],
            "uploaded_at": file_record['uploaded_at'].isoformat() if isinstance(file_record['uploaded_at'], datetime) else file_record['uploaded_at'],
            "total_rules_count": file_record['total_rules_count'],
            "sheets": sheets_list
        }

    async def list_rule_files(self, status='active', limit=50, offset=0): return await self.repository.list_rule_files(status, limit, offset)
    async def get_rule(self, rule_id: str): return await self.repository.get_rule(UUID(rule_id))
    async def update_rule(self, rule_id: str, updates: Dict[str, Any]): return await self.repository.update_rule(UUID(rule_id), updates)
    async def delete_rule(self, rule_id: str, permanent=False): return await self.repository.delete_rule(UUID(rule_id)) if permanent else await self.repository.deactivate_rule(UUID(rule_id))
    async def archive_rule_file(self, file_id: str): return await self.repository.archive_rule_file(UUID(file_id))
    
    async def create_single_rule(self, rule: RuleCreate) -> str:
        """새로운 규칙을 수동으로 생성"""
        rule_data = {
            "rule_file_id": rule.rule_file_id,
            "sheet_name": "Common",
            "row_number": rule.row_number,
            "field_name": rule.column_name,
            "rule_text": rule.rule_text,
            "condition": rule.condition,
            "is_active": True,
            "is_common": rule.is_common,
            "ai_rule_type": rule.ai_rule_type,
            "ai_parameters": rule.ai_parameters
        }
        result = await self.repository.create_single_rule(rule_data)
        return result.get('id')

    async def get_rule_mappings(self, file_id: str):
        file_record = await self.repository.get_rule_file(UUID(file_id))
        if not file_record: return None
        all_rules = await self.repository.get_rules_by_file(UUID(file_id), active_only=True)
        from collections import defaultdict
        groups = defaultdict(list)
        mapped_count = 0; partial_count = 0; unmapped_count = 0
        for r in all_rules:
            conf = r.get('ai_confidence_score', 0) or 0
            if r.get('ai_rule_id'):
                if conf >= 0.8: 
                    status = 'mapped'; mapped_count += 1
                else: 
                    status = 'partial'; partial_count += 1
            else:
                status = 'unmapped'; unmapped_count += 1
            groups[r['field_name']].append({
                "id": str(r['id']), "row_number": r.get('row_number'), "field_name": r.get('field_name'),
                "rule_text": r.get('rule_text'), "condition": r.get('condition'), "mapping_status": status,
                "ai_rule_type": r.get('ai_rule_type'), "ai_parameters": r.get('ai_parameters'),
                "ai_confidence_score": conf, "ai_interpretation_summary": r.get('ai_interpretation_summary')
            })
        fields_list = [{"field_name": name, "rules": sorted(rules, key=lambda x: str(x['row_number'])), "isOpen": False} for name, rules in sorted(groups.items())]
        total = len(all_rules)
        return {
            "file_id": file_id, "file_name": file_record['file_name'],
            "statistics": {
                "total_rules": total, "mapped_count": mapped_count, "partial_count": partial_count, "unmapped_count": unmapped_count,
                "mapping_rate": round((mapped_count + partial_count) / total * 100, 1) if total > 0 else 0
            },
            "sheets": [{"sheet_name": f["field_name"], "rules": f["rules"]} for f in fields_list],
            "available_rule_types": [
                {"value": "required", "label": "필수 입력 (required)"},
                {"value": "allowed_values", "label": "나열형/허용값 (allowed_values)"},
                {"value": "format", "label": "형식 검증 (format)"},
                {"value": "range", "label": "범위 검증 (range)"},
                {"value": "no_duplicates", "label": "중복 금지 (no_duplicates)"},
                {"value": "date_logic", "label": "날짜 논리 (date_logic)"},
                {"value": "composite", "label": "복합 검증 (composite)"},
                {"value": "custom", "label": "사용자 정의 (custom)"}
            ]
        }

    async def reinterpret_rules(self, file_id: str, use_local_parser: bool = True) -> Dict[str, Any]:
        if not self.ai_cache_service: raise Exception("AI Cache Service not initialized")
        return await self.ai_cache_service.interpret_and_cache_rules(file_id, force_reinterpret=True, force_local=use_local_parser)
