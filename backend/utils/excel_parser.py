"""
Excel Parsing Utilities
=======================
Excel 파일(규칙 파일)을 파싱하여 자연어 규칙 리스트를 추출하는 유틸리티 모듈
"""

import io
import re
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook


# =============================================================================
# Composite Rule Detection and Splitting
# =============================================================================

# Keywords that indicate a rule should be split
SPLIT_KEYWORDS = [
    'YYYYMMDD', 'YYYY-MM-DD', 'YYYYMM', 'YYYY/MM/DD',  # Date formats
    '공백', '중복', '필수',  # Common keywords
]

# Comparison operators that indicate field comparison rules
COMPARISON_OPERATORS = ['<=', '>=', '<', '>', '==', '!=', '≤', '≥']


def _is_simple_value_list(text: str) -> bool:
    """
    단순 허용값 목록인지 판별 (분리하지 않음)

    Examples that should NOT be split:
    - "1, 3, 4" (simple numeric list)
    - "1:퇴직, 2:DC전환, 3:사망" (code:description format)
    - "Y, N" (simple choice list)
    - "남, 여" (simple choice list)

    Args:
        text: Rule text to check

    Returns:
        bool: True if it's a simple value list that should not be split
    """
    if not text or not isinstance(text, str):
        return False

    text = text.strip()

    # Check if it's a code:description format (e.g., "1:퇴직, 2:DC전환")
    if ':' in text and ',' in text:
        parts = [p.strip() for p in text.split(',')]
        # If all parts follow pattern "number:text" or "text:text", it's a code list
        code_pattern = re.compile(r'^[\w가-힣]+\s*:\s*[\w가-힣\s]+$')
        if all(code_pattern.match(p) for p in parts if p):
            return True

    # Check if it's a simple numeric list (e.g., "1, 3, 4")
    parts = [p.strip() for p in text.split(',')]
    if all(re.match(r'^-?\d+(\.\d+)?$', p) for p in parts if p):
        return True

    # Check if it's a simple short value list (e.g., "Y, N" or "남, 여")
    # Short means each part is 1-3 characters
    if len(parts) <= 5 and all(len(p) <= 3 for p in parts if p):
        return True

    return False


def _should_split_rule(rule_text: str) -> bool:
    """
    규칙 텍스트가 분리 대상인지 판단

    Split if:
    - Contains multiple distinct rule types separated by comma
    - Contains date format keyword AND field comparison
    - Contains keywords from SPLIT_KEYWORDS mixed with other conditions

    Don't split if:
    - Simple value list (numeric, code:description, Y/N)
    - Single rule type

    Args:
        rule_text: Rule text to check

    Returns:
        bool: True if the rule should be split
    """
    if not rule_text or not isinstance(rule_text, str):
        return False

    text = rule_text.strip()

    # Don't split simple value lists
    if _is_simple_value_list(text):
        return False

    # Check if comma exists
    if ',' not in text:
        return False

    parts = [p.strip() for p in text.split(',')]
    if len(parts) < 2:
        return False

    # Check if any part contains a split keyword
    has_split_keyword = False
    has_comparison = False

    for part in parts:
        # Check for split keywords (date formats, etc.)
        for keyword in SPLIT_KEYWORDS:
            if keyword in part:
                has_split_keyword = True
                break

        # Check for comparison operators (field comparisons)
        for op in COMPARISON_OPERATORS:
            if op in part:
                has_comparison = True
                break

    # Split if we have different types of rules
    # (e.g., date format + field comparison)
    if has_split_keyword and has_comparison:
        return True

    # Split if we have multiple keywords in different parts
    if has_split_keyword:
        keyword_count = 0
        for part in parts:
            for keyword in SPLIT_KEYWORDS:
                if keyword in part:
                    keyword_count += 1
                    break
        # If keyword only appears in one part, and other parts exist, split
        if keyword_count >= 1 and len(parts) > keyword_count:
            # Additional check: other parts should be meaningful (not just whitespace or short)
            non_keyword_parts = []
            for part in parts:
                has_kw = any(kw in part for kw in SPLIT_KEYWORDS)
                if not has_kw and len(part) > 3:
                    non_keyword_parts.append(part)
            if non_keyword_parts:
                return True

    # Split if there are comparison operators
    if has_comparison:
        return True

    return False


def _split_composite_rule_text(rule_text: str) -> List[str]:
    """
    복합 규칙을 개별 조건으로 분리

    Example:
    - "YYYYMMDD, 중간정산기준일<= 입사일" -> ["YYYYMMDD", "중간정산기준일<= 입사일"]
    - "필수, 공백 불가" -> ["필수", "공백 불가"]

    Args:
        rule_text: Composite rule text

    Returns:
        List[str]: List of individual rule texts
    """
    if not rule_text or not isinstance(rule_text, str):
        return [rule_text] if rule_text else []

    text = rule_text.strip()

    # Split by comma
    parts = [p.strip() for p in text.split(',')]

    # Filter out empty parts
    parts = [p for p in parts if p]

    return parts if len(parts) > 1 else [text]


def normalize_sheet_name(name: str) -> str:
    """
    시트 이름 정규화
    - 줄바꿈, 탭 등을 공백으로 치환 (글자 붙음 방지)
    - 연속된 공백을 단일 공백으로 치환

    Args:
        name: 원본 시트 이름

    Returns:
        정규화된 시트 이름
    """
    if not isinstance(name, str):
        return str(name)

    # 제어 문자를 공백으로 치환 (빈 문자열이 아님!)
    normalized = name.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')

    # 연속된 공백 제거
    normalized = re.sub(r'\s+', ' ', normalized)

    return normalized.strip()


def sanitize_sheet_name(name: str) -> str:
    """
    Excel 시트 이름 유효성 처리
    - 최대 31자로 제한
    - Excel에서 허용하지 않는 문자 제거: \ / ? * [ ]

    Args:
        name: 원본 시트 이름

    Returns:
        Excel 호환 시트 이름
    """
    if not name:
        return "Sheet"

    # Excel에서 허용하지 않는 문자 제거
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    sanitized = name
    for char in invalid_chars:
        sanitized = sanitized.replace(char, '')

    # 최대 31자로 제한
    if len(sanitized) > 31:
        sanitized = sanitized[:31]

    return sanitized.strip() or "Sheet"


def get_canonical_name(name: str) -> str:
    """
    비교를 위한 정규화 (모든 공백 제거)

    Args:
        name: 원본 시트 이름

    Returns:
        Canonical 시트 이름 (공백 제거됨)
    """
    norm = normalize_sheet_name(name)
    return "".join(norm.split())


def get_visible_sheet_names(content: bytes) -> List[str]:
    """
    Excel 파일에서 숨겨지지 않은(Visible) 시트 이름 목록만 반환
    - .xlsx: openpyxl로 hidden 시트 제외
    - .xls: pandas로 모든 시트 반환 (hidden 여부 확인 불가)
    
    Args:
        content: Excel 파일의 바이트 내용
        
    Returns:
        List[str]: 숨겨지지 않은 시트 이름 목록 (또는 전체 목록)
    """
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        visible_sheets = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # sheet_state가 'visible'인 경우만 포함 (hidden, veryHidden 제외)
            if sheet.sheet_state == 'visible':
                visible_sheets.append(sheet_name)
        
        wb.close()
        return visible_sheets
    except Exception as e:
        # .xls 파일이거나 손상된 경우 등 openpyxl로 열 수 없을 때
        print(f"[ExcelParser] Warning: Could not check sheet visibility (likely .xls file). Falling back to all sheets. Error: {e}")
        try:
            import pandas as pd
            excel_file = pd.ExcelFile(io.BytesIO(content))
            return excel_file.sheet_names
        except Exception as pd_e:
            print(f"[ExcelParser] Error: Failed to list sheets with pandas: {pd_e}")
            return []


def _detect_reupload_file(ws) -> Tuple[bool, Dict[str, int]]:
    """
    재업로드 파일인지 감지 (이전에 다운로드한 파일)

    첫 번째 행 헤더를 확인하여 재업로드 파일인지 감지합니다.
    재업로드 파일 특징: 헤더에 "AI 규칙 ID", "AI 파라미터", "AI 해석 여부" 등 포함

    Args:
        ws: openpyxl worksheet object

    Returns:
        Tuple[bool, Dict[str, int]]: (is_reupload, column_mapping)
        - is_reupload: True if this is a re-uploaded file
        - column_mapping: Dict mapping column names to their indices (0-based)
    """
    try:
        # Read first row (header)
        header_row = [cell.value for cell in ws[1]]

        # Markers that indicate a re-uploaded file
        reupload_markers = ["AI 규칙 ID", "AI 파라미터", "AI 해석 여부", "AI 규칙 유형"]

        is_reupload = any(marker in header_row for marker in reupload_markers if marker)

        # Build column mapping
        column_mapping = {}
        for idx, header in enumerate(header_row):
            if header:
                column_mapping[str(header)] = idx

        return is_reupload, column_mapping

    except Exception as e:
        print(f"[ExcelParser] Warning: Could not detect reupload file: {e}")
        return False, {}


def parse_rules_from_excel(content: bytes) -> Tuple[List[Dict[str, Any]], Dict[str, int], int, int]:
    """
    Excel B 파일(규칙 파일)을 파싱하여 자연어 규칙 리스트를 반환

    Features:
    - 필드명 기반 규칙 관리 (시트명 제거됨)
    - 복합 규칙을 개별 규칙으로 분리 (예: "YYYYMMDD, 기준일<= 입사일" -> 2개 규칙)
    - 분리된 규칙에 서브 인덱스 부여 (예: row 5 -> "5.1", "5.2")
    - 재업로드 파일 감지 및 처리 (이전 AI 해석 정보 참조용으로 저장)

    Args:
        content: Excel 파일의 바이트 내용

    Returns:
        Tuple containing:
        - natural_language_rules: 파싱된 규칙 리스트 (필드명 기반)
        - field_rule_counts: 필드별 규칙 개수
        - total_raw_rows: 전체 원본 행 수
        - reported_max_row: 엑셀 파일이 메타데이터로 보고하는 총 행 수 (헤더 제외)
    """
    wb = load_workbook(io.BytesIO(content), data_only=True)
    natural_language_rules = []
    field_rule_counts = {}  # 필드별 규칙 개수 (시트 대신)
    total_raw_rows = 0
    reported_max_row = 0

    # 메타데이터 시트 제외 목록 (규칙 매핑에서 제외)
    EXCLUDED_SHEETS = {'파일 정보', '파일정보', 'File Info', 'Metadata', 'metadata', '_metadata'}

    print(f"   [INFO] Found {len(wb.sheetnames)} sheets in rules file: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        # 메타데이터 시트는 건너뛰기
        if sheet_name in EXCLUDED_SHEETS or sheet_name.startswith('_'):
            print(f"   [INFO] Skipping metadata sheet: '{sheet_name}'")
            continue
        ws = wb[sheet_name]
        print(f"   [INFO] Processing rules sheet: '{sheet_name}' (Reported Max Row: {ws.max_row})")

        # 재업로드 파일 감지
        is_reupload, column_mapping = _detect_reupload_file(ws)
        if is_reupload:
            print(f"   [INFO] Detected re-uploaded file format (previous AI interpretation will be stored in note)")

        # 메타데이터 상의 max_row 누적 (헤더 2행 제외)
        if ws.max_row > 2:
            reported_max_row += (ws.max_row - 2)

        consecutive_empty_rows = 0

        for row_idx, row_values in enumerate(ws.iter_rows(min_row=3, max_row=1000, values_only=True), start=3):
            if all(cell is None for cell in row_values):
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= 5:
                    break
                continue

            consecutive_empty_rows = 0
            total_raw_rows += 1

            # Determine column indices based on file format
            if is_reupload and column_mapping:
                # Re-upload file: use column mapping
                column_col = column_mapping.get("컬럼", 2)
                field_col = column_mapping.get("필드명", 3)
                rule_col = column_mapping.get("규칙 내용", 4)
                condition_col = column_mapping.get("조건", 5)
                note_col = column_mapping.get("비고", 6)
                is_common_col = column_mapping.get("공통 여부")
                # AI Fields
                ai_rule_type_col = column_mapping.get("AI 규칙 유형")
                ai_params_col = column_mapping.get("AI 파라미터(JSON)")
                ai_rule_id_col = column_mapping.get("AI 규칙 ID")
                ai_summary_col = column_mapping.get("AI 해석 요약")
                ai_error_col = column_mapping.get("AI 에러 메시지")
            else:
                # Standard file: fixed column positions
                column_col = 2
                field_col = 3
                rule_col = 4
                condition_col = 5
                note_col = 6
                is_common_col = None
                ai_rule_type_col = None
                ai_params_col = None
                ai_rule_id_col = None
                ai_summary_col = None
                ai_error_col = None

            field_name = row_values[field_col] if len(row_values) > field_col else None
            condition = row_values[condition_col] if len(row_values) > condition_col else None
            if condition and "해당없음" in str(condition):
                continue

            column_letter = row_values[column_col] if len(row_values) > column_col else ""
            validation_rule = row_values[rule_col] if len(row_values) > rule_col else ""
            note = row_values[note_col] if len(row_values) > note_col else ""
            safe_field_name = str(field_name) if field_name else "(필드명 없음)"
            rule_text = str(validation_rule) if validation_rule else (f"조건: {condition}" if condition else f"기본 검증 ({safe_field_name})")

            # AI 해석 정보 추출 (파일에 있는 경우)
            prefilled_ai = {}
            if is_reupload:
                if is_common_col is not None and row_values[is_common_col]:
                    prefilled_ai["is_common"] = str(row_values[is_common_col]) == "예"
                if ai_rule_type_col is not None and row_values[ai_rule_type_col]:
                    prefilled_ai["ai_rule_type"] = str(row_values[ai_rule_type_col])
                if ai_params_col is not None and row_values[ai_params_col]:
                    try:
                        import json
                        prefilled_ai["ai_parameters"] = json.loads(str(row_values[ai_params_col]))
                    except: pass
                if ai_rule_id_col is not None and row_values[ai_rule_id_col]:
                    prefilled_ai["ai_rule_id"] = str(row_values[ai_rule_id_col])
                if ai_summary_col is not None and row_values[ai_summary_col]:
                    prefilled_ai["ai_interpretation_summary"] = str(row_values[ai_summary_col])
                if ai_error_col is not None and row_values[ai_error_col]:
                    prefilled_ai["ai_error_message"] = str(row_values[ai_error_col])

            # 필드별 규칙 개수 카운트
            field_rule_counts[safe_field_name] = field_rule_counts.get(safe_field_name, 0) + 1

            # Check if this rule should be split
            if _should_split_rule(rule_text):
                split_rules = _split_composite_rule_text(rule_text)
                for sub_idx, split_rule_text in enumerate(split_rules, start=1):
                    sub_row = f"{row_idx}.{sub_idx}"
                    rule_entry = {
                        "row": sub_row,
                        "column_letter": str(column_letter) if column_letter else "",
                        "field": safe_field_name,
                        "rule_text": split_rule_text,
                        "condition": str(condition) if condition else "",
                        "note": str(note) if note else "",
                        "prefilled_ai": prefilled_ai # 기해석 정보 포함
                    }
                    natural_language_rules.append(rule_entry)
            else:
                rule_entry = {
                    "row": str(row_idx),
                    "column_letter": str(column_letter) if column_letter else "",
                    "field": safe_field_name,
                    "rule_text": rule_text,
                    "condition": str(condition) if condition else "",
                    "note": str(note) if note else "",
                    "prefilled_ai": prefilled_ai # 기해석 정보 포함
                }
                natural_language_rules.append(rule_entry)

    return natural_language_rules, field_rule_counts, total_raw_rows, reported_max_row
