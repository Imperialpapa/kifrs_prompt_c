"""
K-IFRS 1019 DBO Validation System - FastAPI Backend
===================================================
전체 시스템을 통합하는 웹 API

Endpoints:
- POST /upload: 파일 업로드
- POST /validate: 검증 실행
- GET /health: 헬스체크
"""

from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import pandas as pd
import io
from typing import List, Dict, Any
import traceback
from datetime import datetime

from models import (
    ValidationResponse,
    AIInterpretationResponse,
    RuleConflict,
    ValidationErrorGroup
)
from ai_layer import AIRuleInterpreter
from rule_engine import RuleEngine

# =============================================================================
# FastAPI 앱 초기화
# =============================================================================

app = FastAPI(
    title="K-IFRS 1019 DBO Validation System",
    description="AI-Powered Data Validation for Defined Benefit Obligations",
    version="1.0.0"
)

# CORS 설정 (모바일에서 접근 가능하도록)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 프로덕션에서는 구체적인 도메인 지정
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =============================================================================
# 전역 상태
# =============================================================================

ai_interpreter = AIRuleInterpreter()


# =============================================================================
# 헬퍼 함수
# =============================================================================

def parse_rules_from_excel(content: bytes) -> tuple[List[Dict[str, Any]], Dict[str, int], int, int]:
    """
    Excel B 파일(규칙 파일)을 파싱하여 자연어 규칙 리스트를 반환
    
    Returns:
        (natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row)
        - reported_max_row: 엑셀 파일이 메타데이터로 보고하는 총 행 수 (헤더 제외)
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(content), data_only=True)
    natural_language_rules = []
    sheet_row_counts = {}
    total_raw_rows = 0
    reported_max_row = 0
    
    print(f"   [INFO] Found {len(wb.sheetnames)} sheets in rules file: {wb.sheetnames}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"   [INFO] Processing rules sheet: '{sheet_name}' (Reported Max Row: {ws.max_row})")
        
        # 메타데이터 상의 max_row 누적 (헤더 2행 제외)
        if ws.max_row > 2:
            reported_max_row += (ws.max_row - 2)
        
        last_sheet_name_val = None 
        consecutive_empty_rows = 0
        
        for row_idx, row_values in enumerate(ws.iter_rows(min_row=3, max_row=1000, values_only=True), start=3):
            if all(cell is None for cell in row_values):
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= 5: 
                    break
                continue
            
            consecutive_empty_rows = 0
            total_raw_rows += 1
            
            raw_sheet_val = row_values[1] if len(row_values) > 1 else None
            if isinstance(raw_sheet_val, str) and not raw_sheet_val.strip():
                raw_sheet_val = None
            if raw_sheet_val is not None:
                last_sheet_name_val = raw_sheet_val
            
            current_sheet_name = last_sheet_name_val
            
            if current_sheet_name:
                disp_name = normalize_sheet_name(str(current_sheet_name))
                sheet_row_counts[disp_name] = sheet_row_counts.get(disp_name, 0) + 1
            
            if not current_sheet_name:
                continue
                
            field_name = row_values[3] if len(row_values) > 3 else None
            condition = row_values[5] if len(row_values) > 5 else None
            if condition and "해당없음" in str(condition):
                continue
            
            column_letter = row_values[2] if len(row_values) > 2 else ""
            validation_rule = row_values[4] if len(row_values) > 4 else ""
            note = row_values[6] if len(row_values) > 6 else ""
            safe_field_name = str(field_name) if field_name else "(필드명 없음)"
            rule_text = str(validation_rule) if validation_rule else (f"조건: {condition}" if condition else f"기본 검증 ({safe_field_name})")
            
            rule_entry = {
                "sheet": get_canonical_name(str(current_sheet_name)),
                "display_sheet_name": normalize_sheet_name(str(current_sheet_name)),
                "row": row_idx,
                "column_letter": str(column_letter) if column_letter else "",
                "field": safe_field_name,
                "rule_text": rule_text,
                "condition": str(condition) if condition else "",
                "note": str(note) if note else ""
            }
            natural_language_rules.append(rule_entry)
            
    return natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row

def normalize_sheet_name(name: str) -> str:
    """
    시트 이름 정규화
    - 줄바꿈, 탭 등을 공백으로 치환 (글자 붙음 방지)
    - 연속된 공백을 단일 공백으로 치환
    """
    if not isinstance(name, str):
        return str(name)
    
    # 제어 문자를 공백으로 치환 (빈 문자열이 아님!)
    normalized = name.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    
    # 연속된 공백 제거
    import re
    normalized = re.sub(r'\s+', ' ', normalized)
    
    return normalized.strip()

def get_canonical_name(name: str) -> str:
    """
    비교를 위한 정규화 (모든 공백 제거)
    """
    norm = normalize_sheet_name(name)
    return "".join(norm.split())

def group_errors(errors: list) -> List[ValidationErrorGroup]:
    """
    동일한 인지 내용을 그룹화하여 집계

    Args:
        errors: ValidationError 리스트

    Returns:
        List[ValidationErrorGroup]: 그룹화된 인지 목록
    """
    from collections import defaultdict

    # (시트, 컬럼, 규칙ID, 메시지)를 키로 그룹화
    groups = defaultdict(list)

    for error in errors:
        key = (
            error.sheet or "",
            error.column,
            error.rule_id,
            error.message
        )
        groups[key].append(error)

    # ValidationErrorGroup 객체 생성
    error_groups = []
    for (sheet, column, rule_id, message), error_list in groups.items():
        # 행 번호 수집
        affected_rows = [e.row for e in error_list]

        # 샘플 값 수집 (최대 3개, 중복 제거)
        sample_values = []
        seen_values = set()
        for e in error_list:
            val_str = str(e.actual_value)
            if val_str not in seen_values and len(sample_values) < 3:
                sample_values.append(e.actual_value)
                seen_values.add(val_str)

        error_group = ValidationErrorGroup(
            sheet=sheet,
            column=column,
            rule_id=rule_id,
            message=message,
            affected_rows=sorted(affected_rows),
            count=len(error_list),
            sample_values=sample_values,
            expected=error_list[0].expected if error_list else None,
            source_rule=error_list[0].source_rule if error_list else ""
        )
        error_groups.append(error_group)

    # 인지 개수 많은 순으로 정렬
    error_groups.sort(key=lambda x: x.count, reverse=True)

    return error_groups


# =============================================================================
# API Endpoints
# =============================================================================

@app.get("/")
async def root():
    """
    루트 엔드포인트
    """
    return {
        "service": "K-IFRS 1019 DBO Validation System",
        "version": "1.2.5",
        "build_date": "2025-01-13 10:00:00",
        "status": "operational",
        "features": [
            "다중 시트 검증 지원 (자동 매핑)",
            "인지 항목 집계",
            "동적 AI 규칙 생성"
        ],
        "endpoints": {
            "health": "/health",
            "validate": "/validate (POST)",
            "version": "/version (GET)"
        }
    }


@app.get("/health")
async def health_check():
    """
    헬스체크
    """
    return {
        "status": "healthy",
        "ai_layer": "operational",
        "rule_engine": "operational"
    }


@app.get("/version")
async def get_version():
    """버전 정보 조회"""
    import sys
    import platform
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return {
        "system_version": "1.2.6",
        "build_date": now_str,
        "build_time": now_str,
        "server_start_time": now_str,
        "python_version": sys.version,
        "platform": platform.platform(),
        "features": {
            "multi_sheet_validation": True,
            "error_grouping": True,
            "dynamic_rule_generation": True,
            "excel_download": True,
            "fuzzy_sheet_matching": True,
            "sheet_matching_stats": True
        }
    }


@app.post("/validate", response_model=ValidationResponse)
async def validate_data(
    employee_file: UploadFile = File(..., description="직원 데이터 파일 (Excel A)"),
    rules_file: UploadFile = File(..., description="검증 규칙 파일 (Excel B)")
):
    """
    데이터 검증 실행
    
    Process:
    1. Excel A (직원 데이터) 읽기
    2. Excel B (자연어 규칙) 읽기
    3. AI에게 규칙 해석 요청
    4. 결정론적 엔진으로 검증 실행
    5. 결과 반환
    """
    try:
        # =====================================================================
        # Step 1: Excel A 읽기 (직원 데이터) - 다중 시트 지원
        # =====================================================================
        print("[Step 1] Reading employee data...")
        employee_content = await employee_file.read()

        # 모든 시트 읽기
        excel_file = pd.ExcelFile(io.BytesIO(employee_content))
        
        # 통합 데이터 맵: Canonical Name -> { display_name, original_name, df }
        # 이것이 시트 데이터 조회 및 매칭의 Single Source of Truth가 됩니다.
        sheet_data_map = {}
        sheet_mapping_info = {} # Debug info for logging

        print(f"   [INFO] Raw sheet names in file: {excel_file.sheet_names}")

        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(io.BytesIO(employee_content), sheet_name=sheet_name)
            
            # 이름 변환
            norm_name = normalize_sheet_name(sheet_name)
            canonical_name = get_canonical_name(sheet_name)
            
            sheet_data_map[canonical_name] = {
                "display_name": norm_name,
                "original_name": sheet_name,
                "df": df
            }
            sheet_mapping_info[canonical_name] = sheet_name
            
            print(f"   [OK] Loaded sheet '{sheet_name}' -> canonical '{canonical_name}': {len(df)} rows")

        print(f"[OK] Loaded {len(sheet_data_map)} sheets from employee file")
        
        # =====================================================================
        # Step 2: Excel B 읽기 (자연어 규칙)
        # =====================================================================
        print("\n[Step 2] Reading validation rules...")
        rules_content = await rules_file.read()

        # 헬퍼 함수를 사용하여 규칙 파싱 (로직 통합)
        natural_language_rules, sheet_row_counts, total_raw_rows, reported_max_row = parse_rules_from_excel(rules_content)

        # 디버깅: 시트별 규칙 개수 출력
        from collections import Counter
        sheet_counts = Counter(r['display_sheet_name'] for r in natural_language_rules)
        print(f"[OK] Loaded {len(natural_language_rules)} validation rules from {len(sheet_counts)} sheets:")
        for sheet, count in sheet_counts.items():
            print(f"   - '{sheet}': {count} rules")

        # [수정] 매칭 통계 계산을 위한 리스트 미리 정의 (Step 5에서 사용)
        # 필터링 전의 전체 시트 목록(sheet_row_counts)을 기준으로 합니다.
        all_rule_sheets_display_unfiltered = sorted(list(sheet_row_counts.keys()))
        all_rule_sheets_canonical_unfiltered = [get_canonical_name(name) for name in all_rule_sheets_display_unfiltered]

        # =====================================================================
        # Step 3: AI 규칙 해석
        # =====================================================================
        print("\n[Step 3] AI interpreting rules...")
        ai_response: AIInterpretationResponse = await ai_interpreter.interpret_rules(
            natural_language_rules
        )

        print(f"[OK] AI interpreted {len(ai_response.rules)} rules")
        print(f"[WARN] Detected {len(ai_response.conflicts)} conflicts")

        # =====================================================================
        # Step 4: 결정론적 검증 실행 - 시트별 처리 (통합 매칭 로직 적용)
        # =====================================================================
        print("\n[Step 4] Running deterministic validation...")
        engine = RuleEngine()
        all_errors = []
        all_sheets_summary = {}

        # 규칙을 Canonical Name 기준으로 미리 그룹화 (O(1) 조회를 위해)
        from collections import defaultdict
        rules_by_sheet = defaultdict(list)
        for rule in ai_response.rules:
            rules_by_sheet[rule.source.sheet_name].append(rule)

        # 시트별 데이터 순회 (Canonical Key 사용)
        for canonical_name, data in sheet_data_map.items():
            display_name = data["display_name"]
            df = data["df"]
            
            print(f"\n   [SHEET] Validating sheet: '{display_name}' (Canonical: '{canonical_name}')")
            
            # 규칙 조회
            sheet_rules = rules_by_sheet.get(canonical_name, [])
            
            if not sheet_rules:
                print(f"   [WARN] No rules found for sheet: '{display_name}', skipping...")
                continue
            
            # 검증 실행
            errors = engine.validate(df, sheet_rules)
            summary = engine.get_summary(len(df), len(sheet_rules))

            # 시트 정보를 에러에 추가
            for error in errors:
                error.sheet = display_name

            all_errors.extend(errors)
            
            # 요약 정보 저장
            all_sheets_summary[display_name] = {
                "total_rows": len(df),
                "error_rows": summary.error_rows,
                "valid_rows": summary.valid_rows,
                "total_errors": len(errors),
                "rules_applied": len(sheet_rules)
            }

        # 전체 요약 계산
        total_rows = sum(s["total_rows"] for s in all_sheets_summary.values())
        total_error_rows = sum(s["error_rows"] for s in all_sheets_summary.values())

        # =====================================================================
        # Step 5: 응답 생성
        # =====================================================================
        # 통합 요약 생성
        from models import ValidationSummary
        combined_summary = ValidationSummary(
            total_rows=total_rows,
            valid_rows=total_rows - total_error_rows,
            error_rows=total_error_rows,
            total_errors=len(all_errors),
            rules_applied=len(ai_response.rules),
            timestamp=datetime.now()
        )

        # 인지 내용 집계
        error_groups = group_errors(all_errors)

        # 매칭 통계 계산 (통합된 자료구조 활용)
        # AI 규칙 개수 집계 (시트별 - Canonical Name 기준)
        from collections import Counter
        rule_counts_by_canonical = Counter(rule.source.sheet_name for rule in ai_response.rules)

        # 매칭 연산
        data_sheets_set = set(sheet_data_map.keys())
        rule_sheets_set = set(all_rule_sheets_canonical_unfiltered)
        
        matched_sheets_set = data_sheets_set.intersection(rule_sheets_set)
        unmatched_sheets_set = rule_sheets_set - data_sheets_set
        
        matched_sheets_count = len(matched_sheets_set)
        
        # 매칭 안 된 시트들의 Display Name 찾기
        unmatched_sheet_names = []
        for c_name in unmatched_sheets_set:
            d_name = next((d for d in all_rule_sheets_display_unfiltered if get_canonical_name(d) == c_name), c_name)
            unmatched_sheet_names.append(d_name)

        all_data_sheets = sorted(list(sheet_mapping_info.values()))

        # 드롭다운 표시용 문자열 생성: "시트명 (N행 / M규칙)"
        display_list = []
        for d_name in all_rule_sheets_display_unfiltered:
            c_name = get_canonical_name(d_name)
            raw_rows_in_sheet = sheet_row_counts.get(d_name, 0)
            rule_count_in_sheet = rule_counts_by_canonical.get(c_name, 0)
            display_list.append(f"{d_name} ({raw_rows_in_sheet}행 / {rule_count_in_sheet}규칙)")

        matching_stats = {
            "total_rule_sheets": len(all_rule_sheets_canonical_unfiltered),
            "matched_sheets": matched_sheets_count,
            "unmatched_sheet_names": unmatched_sheet_names,
            "all_rule_sheets": display_list,
            "all_data_sheets": all_data_sheets,
            # [DEBUG CODE] 사용자 요청에 의한 디버깅용 필드 (추후 삭제 가능)
            "total_raw_rows": total_raw_rows, 
            "reported_max_row": reported_max_row,
            "total_rules_count": len(ai_response.rules)
            # [END DEBUG CODE]
        }

        response = ValidationResponse(
            validation_status="PASS" if len(all_errors) == 0 else "FAIL",
            summary=combined_summary,
            errors=all_errors[:200],  # 최대 200개까지만 반환 (대량 인지 대응)
            error_groups=error_groups,  # 집계된 인지 목록
            conflicts=ai_response.conflicts,
            rules_applied=ai_response.rules,
            metadata={
                "employee_file_name": employee_file.filename,
                "rules_file_name": rules_file.filename,
                "ai_model_version": "claude-sonnet-4-20250514",
                "system_version": "1.0.0",
                "ai_processing_time_seconds": ai_response.processing_time_seconds,
                "total_errors": len(all_errors),
                "errors_shown": min(len(all_errors), 200),
                "error_groups_count": len(error_groups),
                "sheets_summary": all_sheets_summary,
                "matching_stats": matching_stats
            }
        )

        print("\n[OK] Response ready")
        return response

    except Exception as e:
        print(f"\n[ERROR] Error: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail={
                "error": "Validation failed",
                "message": str(e),
                "type": type(e).__name__
            }
        )


@app.post("/interpret-rules")
async def interpret_rules_only(
    rules_file: UploadFile = File(..., description="검증 규칙 파일 (Excel B)")
):
    """
    규칙만 해석 (검증 실행 없이)
    - 디버깅 및 규칙 검토용
    """
    try:
        # Excel B 읽기
        rules_content = await rules_file.read()
        
        # 헬퍼 함수 사용하여 규칙 파싱 (로직 통합)
        natural_language_rules, _, _, _ = parse_rules_from_excel(rules_content)
        
        # AI 해석
        ai_response = await ai_interpreter.interpret_rules(natural_language_rules)
        
        return {
            "status": "success",
            "rules_count": len(ai_response.rules),
            "conflicts_count": len(ai_response.conflicts),
            "rules": [rule.dict() for rule in ai_response.rules],
            "conflicts": [conflict.dict() for conflict in ai_response.conflicts],
            "summary": ai_response.ai_summary
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail={
                "error": "Rule interpretation failed",
                "message": str(e)
            }
        )


@app.get("/kifrs-references")
async def get_kifrs_references():
    """
    K-IFRS 1019 참조 정보 조회
    """
    from models import KIFRS_1019_REFERENCES
    return KIFRS_1019_REFERENCES


@app.post("/download-results")
async def download_validation_results(validation_response: ValidationResponse):
    """
    검증 결과를 Excel 파일로 다운로드

    Args:
        validation_response: 검증 결과 데이터

    Returns:
        Excel 파일 (StreamingResponse)
    """
    try:
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. 요약 시트
            summary_data = {
                "항목": ["검증 상태", "전체 행 수", "정상 행 수", "오류 행 수", "총 오류 수", "적용된 규칙 수", "검증 시각"],
                "값": [
                    validation_response.validation_status,
                    validation_response.summary.total_rows,
                    validation_response.summary.valid_rows,
                    validation_response.summary.error_rows,
                    validation_response.summary.total_errors,
                    validation_response.summary.rules_applied,
                    validation_response.summary.timestamp.strftime("%Y-%m-%d %H:%M:%S")
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name="검증 요약", index=False)

            # 2. 인지 항목 집계 시트
            if validation_response.error_groups:
                groups_data = []
                for group in validation_response.error_groups:
                    rows_str = ', '.join(map(str, group.affected_rows[:20]))
                    if len(group.affected_rows) > 20:
                        rows_str += f" 외 {len(group.affected_rows) - 20}개"

                    groups_data.append({
                        "시트명": group.sheet,
                        "열": group.column,
                        "규칙ID": group.rule_id,
                        "인지 메시지": group.message,
                        "인지 횟수": group.count,
                        "영향받은 행": rows_str,
                        "샘플 값": ", ".join(map(str, group.sample_values)),
                        "예상 값": group.expected or "",
                        "원본 규칙": group.source_rule
                    })
                df_groups = pd.DataFrame(groups_data)
                df_groups.to_excel(writer, sheet_name="인지 항목 집계", index=False)

            # 3. 개별 인지 목록 시트
            if validation_response.errors:
                errors_data = []
                for error in validation_response.errors:
                    errors_data.append({
                        "시트명": error.sheet or "",
                        "행": error.row,
                        "열": error.column,
                        "규칙ID": error.rule_id,
                        "인지 메시지": error.message,
                        "실제 값": str(error.actual_value),
                        "예상 값": error.expected or "",
                        "원본 규칙": error.source_rule
                    })
                df_errors = pd.DataFrame(errors_data)
                df_errors.to_excel(writer, sheet_name="개별 인지 목록", index=False)

            # 4. 규칙 충돌 시트
            if validation_response.conflicts:
                conflicts_data = []
                for conflict in validation_response.conflicts:
                    conflicts_data.append({
                        "규칙ID": conflict.rule_id,
                        "충돌 유형": conflict.conflict_type,
                        "설명": conflict.description,
                        "K-IFRS 1019 참조": conflict.kifrs_reference or "",
                        "영향받는 규칙": ", ".join(conflict.affected_rules),
                        "권장사항": conflict.recommendation,
                        "심각도": conflict.severity
                    })
                df_conflicts = pd.DataFrame(conflicts_data)
                df_conflicts.to_excel(writer, sheet_name="규칙 충돌", index=False)

            # 5. 적용된 규칙 시트
            if validation_response.rules_applied:
                rules_data = []
                for rule in validation_response.rules_applied:
                    rules_data.append({
                        "규칙ID": rule.rule_id,
                        "시트명": rule.source.sheet_name,
                        "필드명": rule.field_name,
                        "규칙 유형": rule.rule_type,
                        "파라미터": str(rule.parameters),
                        "오류 메시지": rule.error_message_template,
                        "원본 규칙": rule.source.original_text,
                        "신뢰도": rule.confidence_score
                    })
                df_rules = pd.DataFrame(rules_data)
                df_rules.to_excel(writer, sheet_name="적용된 규칙", index=False)

        output.seek(0)

        # 파일명 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"DBO_Validation_Results_{timestamp}.xlsx"

        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail={
                "error": "Failed to generate Excel file",
                "message": str(e)
            }
        )


# =============================================================================
# 예외 핸들러
# =============================================================================

@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """
    전역 예외 핸들러
    """
    return JSONResponse(
        status_code=500,
        content={
            "error": "Internal server error",
            "message": str(exc),
            "type": type(exc).__name__
        }
    )


# =============================================================================
# 서버 실행 (개발용)
# =============================================================================

if __name__ == "__main__":
    import uvicorn

    print("""
    =================================================================
      K-IFRS 1019 DBO Validation System
      AI-Powered Data Validation for Defined Benefit Obligations
    =================================================================

    Starting server...
    Mobile-optimized UI available at: http://localhost:8000
    API Documentation: http://localhost:8000/docs

    """)
    
    uvicorn.run(
        "main:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
        log_level="info"
    )